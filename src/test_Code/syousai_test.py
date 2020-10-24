import openpyxl
import app_ma,es.py
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("test01.xlsx")
wb_s = openpyxl.load_workbook("詳細計画_Sample_20201016.xlsm")
wb_bps = openpyxl.load_workbook("進捗報告ファイル_20201016.xlsm")

b = wb_bps.sheetnames

print(wb_s.sheetnames)
print(b)

'''
& c:/works/venv_Projects/openpyxl_PRJ/Scripts/pythonw.exe 
c:/works/venv_Projects/openpyxl_PRJ/src/test_Code/syousai_test.py
'''

DPC_SP_101

for row in range(10:1000):
    column_letter = get_column_letter(row)
    