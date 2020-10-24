import openpyxl as xp

wb = xp.load_workbook("kinou.xlsx")
wb2 = xp.load_workbook("kinou.xlsx", data_only=True)

#関数ごとコピー
def copy_and_paste(sheetname, num_min_row, num_max_row, num_min_column, num_max_column, sheetname_paste):
    """
    docstring
    """
    ws = wb[str(sheetname)]

    paste_ws = wb[sheetname_paste]

    for j in range(num_min_column, num_max_column):
        for i in range(num_min_row, num_max_row):
            copy = ws.cell(row=i, column=j).value
            print("セル" + str(i) + "," + str(j) + ":" + str(copy))
            paste_ws.cell(row=i, column=j, value=copy)
            i += 1
        j += 1

#値のみコピー
def copy_and_paste2(sheetname, num_min_row, num_max_row, num_min_column, num_max_column, sheetname_paste):
    """
    docstring
    """
    ws2 = wb2[str(sheetname)]

    paste_ws = wb[sheetname_paste]

    for j in range(num_min_column, num_max_column):
        for i in range(num_min_row, num_max_row):
            copy = ws2.cell(row=i, column=j).value
            print("セル" + str(i) + "," + str(j) + ":" + str(copy))
            paste_ws.cell(row=i, column=j, value=copy)
            i += 1
        j += 1


#copy_and_paste("Sheet1", 3, 27, 2, 6)
copy_and_paste("Sheet2", 1, 15, 1, 6, "Sheet3")

new_ws = wb.create_sheet(title="Sheet4")
copy_and_paste2("Sheet2", 1, 15, 1, 6, "Sheet4")

wb.save("kinou3.xlsx")
