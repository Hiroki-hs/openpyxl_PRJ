import openpyxl

#----------------------------------------------------------------------------------------
# ブックシートを追加し、B~E列を新規シートにコピー
#----------------------------------------------------------------------------------------

#ブックの読みこみ
wb = openpyxl.load_workbook("kinou.xlsx")

#新規シートを作成。インデックス（第2引数）に0を指定すると先頭にシートを追加する。
ws = wb.create_sheet("pasteField",0)

#シート名を全部表示
print("-------------シート名-------------")
print(wb.sheetnames)

#コピー元のシートをアクティブ化
ws_active = wb["Sheet1"]
ws_paste = wb["pasteField"]

#指定範囲のセルをコピー
'''
for i in ws_active.iter_rows(min_row=3, min_col=2, max_row=26, max_col=5):
    for j in i:
        print(j.value)
'''

print("----------------------コピー開始----------------------")
for columns in range(2, 6):   
    for i in range(3,27):
        copy = ws_active.cell(row=i, column=columns).value
        print("次の値をコピーします:" + str(copy))
        ws_paste.cell(row=i-2, column=columns, value=copy)
        i += 1
    columns += 1

    
#ワークブックを上書き保存
wb.save("kinou2.xlsx")

print("----------------------コピー完了----------------------")