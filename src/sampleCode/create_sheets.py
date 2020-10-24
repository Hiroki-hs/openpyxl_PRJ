import openpyxl


#Excelファイルを開く
wb = openpyxl.load_workbook("test01.xlsx")

#新規シートを追加
ws = wb.create_sheet(title="sample_sheet_02")
s_names = wb.sheetnames
print(s_names)

#シートのコピー
copy_wb = wb.copy_worksheet(wb["test_sheet_01"])
s_names = wb.sheetnames
print(s_names)

wb.close()
