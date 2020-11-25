import openpyxl
import pprint

path = "D:\\sample_datas\\Shiseido_financialStatement.xlsx"
path2 = "D:\\sample_datas\\Shiseido_financialStatement2.xlsx"

wb = openpyxl.load_workbook(path)
ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]

#print(ws1.title)
#print(ws2.title)

set1 = set()
for i in range(25):
    bs = ws1.cell(row=i+1, column=2).value
    set1.add(bs)

#print(set1)

set2 = set()
for i in range(25):
    bs = ws2.cell(row=i+1, column=2).value
    set2.add(bs)

#print(set2)

set3 = set2 - set1
print(set3)

delete = {}
for i in set3:
    for j in range(1,26):
        target = ws2.cell(row = j + 1, column=2).value
        if target == i:
            delete[ws2.cell(row = j + 1, column=2).value] = j+1
        
pprint.pprint(delete)
sort_del = sorted(delete.values())
pprint.pprint("ソート後:{}".format(sort_del))
    
ws3 = wb.create_sheet("delete_file")
cnt=2
del_cnt = 0

for q in sort_del:
    for k in range(1,7):
        ws3.cell(row=cnt , column = k, value=ws2.cell(row = q - del_cnt, column = k).value)
    ws2.delete_rows(q - del_cnt)
    cnt+=1
    del_cnt += 1
#print(del_cnt)

wb.save(path2)